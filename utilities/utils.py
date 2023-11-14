import openpyxl


def get_data_from_excel(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    row_count = ws.max_row
    column_count = ws.max_column
    data_list = []
    for rows in range(2, row_count+1):
        row = []
        for columns in range(1, column_count+1):
            row.append(ws.cell(row=rows, column=columns).value)
        data_list.append(row)
    return data_list




